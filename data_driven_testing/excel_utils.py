import os
import openpyxl
from openpyxl.styles import PatternFill  # Import for coloring cells


def fill_cell_color(file_path: str, sheet_name: str, row: int, col: int, color: str):
    colors = {
        'red': PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid"),
        'light_red': PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),  # Light red
        'blue': PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid"),  # Light blue
        'light_green': PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),  # Light green
        'green': PatternFill(start_color="60b212", end_color="60b212", fill_type="solid"),
        'yellow': PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),  # Light yellow
        'Grey': PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"),  # Grey
        'purple': PatternFill(start_color="E6CCFF", end_color="E6CCFF", fill_type="solid")  # Light purple
    }
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row, col).fill = colors.get(color)
    workbook.save(file_path)


def write_to_cell(file_path: str, sheet_name: str, row: int, col: int, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row, col).value = data
    workbook.save(file_path)


def read_data(file_path: str, sheet_name: str, row: int, col: int):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.cell(row, col).value


def get_row_count(file_path: str, sheet_name: str) -> int:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_col_count(file_path: str, sheet_name: str) -> int:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_sheet(file_path: str) -> None:
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook['Sheet1']
        rows = sheet.max_row
        cols = sheet.max_column

        print(f"Number of rows: {rows}")
        print(f"Number of cols: {cols}")

        # how to access a cell -> print(sheet[2][2].value)
        # note: the cells start from [1][1] not from [0][0]

        # print the sheet
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                print(f'[{sheet.cell(r, c).value}]', end='\t\t')
            print(end='\n')

    except FileNotFoundError as e:
        print(e)
    except Exception as e:
        print(f"An error occurred: {e}")
# ------------------------------------------------------------------------------------


def write_to_sheet(file_path: str, rows: int, cols: int) -> None:
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        # Load the workbook from the file
        workbook = openpyxl.load_workbook(file_path)

        sheet = workbook.active  # OR sheet = workbook["SHEET_NAME"]

        # Note: The first row or column integer is 1, not 0.
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                sheet.cell(r, c).value = (r + c)  # write diff values to the cells

        # DIRECT sheet modification:
        # sheet.cell(1, 1).value = 123
        # sheet.cell(1, 2).value = "smith"
        # sheet.cell(1, 3).value = "engineer"
        #
        # sheet.cell(2, 1).value = 567
        # sheet.cell(2, 2).value = "john"
        # sheet.cell(2, 3).value = "manager"
        #
        # sheet.cell(3, 1).value = 567
        # sheet.cell(3, 2).value = "david"
        # sheet.cell(3, 3).value = "developer"

        # NOTE: if the written data is not SAVED it will be gone
        workbook.save(file_path)

    except FileNotFoundError as e:
        print(e)
    except Exception as e:
        print(f"An error occurred: {e}")
# ------------------------------------------------------------------------------------


def apply_row_color(file_path: str, rows: int, cols: int) -> None:
    """
    Apply coloring to cells in a row based on its parity (even or odd).
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        # Load the workbook from the file
        workbook = openpyxl.load_workbook(file_path)

        sheet = workbook.active  # OR sheet = workbook["SHEET_NAME"]

        # Define the color fills
        red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
        blue = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")  # Light blue
        green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light green
        yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light yellow
        Grey = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Grey
        purple = PatternFill(start_color="E6CCFF", end_color="E6CCFF", fill_type="solid")  # Light purple

        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                if r % 2 == 0:
                    sheet.cell(r, c).fill = yellow
                else:
                    sheet.cell(r, c).fill = blue

        workbook.save(file_path)

    except FileNotFoundError as e:
        print(e)
    except Exception as e:
        print(f"An error occurred: {e}")

# ------------------------------------------------------------------------------------


def create_valid_excel_file(file_path: str):
    """
    Create a valid Excel file at the specified path.
    """
    try:
        # Create a new workbook
        workbook = openpyxl.Workbook()

        # Access the default active sheet
        sheet = workbook.active
        sheet.title = "Sheet1"  # Name the sheet (optional)

        # Add some sample data to ensure the file is valid
        sheet.cell(1, 1, value="Sample Data")
        sheet.cell(2, 1, value="Another Row")

        # Save the workbook to the specified file path
        workbook.save(file_path)
        print(f"Excel file created successfully at: {file_path}")

    except Exception as e:
        print(f"An error occurred: {e}")
# ------------------------------------------------------------------------------------


if __name__ == '__main__':
    file_path_1 = r'D:/data/file1.xlsx'  # Use forward slashes to avoid escape sequence issues
    file_path_2 = r'D:/data/file2.xlsx'

    read_sheet(file_path_1)
    create_valid_excel_file(file_path_2)
    write_to_sheet(file_path_2, 10, 15)
    apply_row_color(file_path_2, 10, 15)


